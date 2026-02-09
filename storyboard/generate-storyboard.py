"""
MPP Portal eLearning - Complete Storyboard Generator

Generates an 11-sheet Excel workbook from JSON data files and screenshots.
Each sheet covers one section of the MPP Portal with embedded screenshots,
modal text, voiceover scripts, and quiz questions.

Usage: python storyboard/generate-storyboard.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import json
import os
import glob

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, 'data')
SCREENSHOTS_DIR = os.path.join(SCRIPT_DIR, 'screenshots')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, 'MPP_Portal_eLearning_Storyboard.xlsx')

# ============================================================
# Styles
# ============================================================
NAVY_FILL = PatternFill('solid', fgColor='0C1F5B')
LIGHT_BLUE_FILL = PatternFill('solid', fgColor='E8F4FD')
QUIZ_FILL = PatternFill('solid', fgColor='FFF3CD')
CORRECT_FILL = PatternFill('solid', fgColor='D4EDDA')
INCORRECT_FILL = PatternFill('solid', fgColor='F8D7DA')

WHITE_FONT = Font(bold=True, color='FFFFFF', size=12, name='Calibri')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
BODY_FONT = Font(size=10, name='Calibri')
QUIZ_FONT = Font(size=10, name='Calibri', bold=True)
STEP_FONT = Font(size=14, bold=True, name='Calibri')

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

WRAP_ALIGN = Alignment(wrap_text=True, vertical='top', horizontal='left')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Section JSON files in order
SECTION_FILES = [
    '01-dashboard.json',
    '02-protege-signup.json',
    '03-agreement-overview.json',
    '04-agreement-team.json',
    '05-mentor-company-info.json',
    '06-protege-company-info.json',
    '07-needs-assessment.json',
    '08-terms-of-agreement.json',
    '09-dap.json',
    '10-white-paper.json',
    '11-mpa-proposal.json',
]


def setup_column_widths(ws):
    """Set column widths: Step#, Screenshot, Modal Text, Voiceover, Notes."""
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 58
    ws.column_dimensions['E'].width = 28


def add_headers(ws):
    """Add styled header row."""
    headers = ['Step #', 'Screenshot', 'Modal Text', 'Voiceover Script', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28


def add_tour_step(ws, row, step):
    """Add a single tour step row with embedded screenshot."""
    ws.row_dimensions[row].height = 180

    # Column A: Step #
    cell_a = ws.cell(row=row, column=1, value=step.get('stepId', ''))
    cell_a.font = STEP_FONT
    cell_a.alignment = CENTER_ALIGN
    cell_a.border = THIN_BORDER

    # Column B: Screenshot
    cell_b = ws.cell(row=row, column=2)
    img_path = os.path.join(SCREENSHOTS_DIR, step.get('screenshotFile', ''))
    if os.path.exists(img_path):
        try:
            img = XLImage(img_path)
            max_width = 320
            max_height = 170
            if img.width > 0 and img.height > 0:
                aspect = img.width / img.height
                if img.width > max_width:
                    img.width = max_width
                    img.height = int(max_width / aspect)
                if img.height > max_height:
                    img.height = max_height
                    img.width = int(max_height * aspect)
            ws.add_image(img, f'B{row}')
        except Exception as e:
            cell_b.value = f"[{step.get('screenshotFile', 'missing')}]"
            print(f"  [WARN] Image error: {e}")
    else:
        cell_b.value = f"[{step.get('screenshotFile', 'missing')}]"
    cell_b.alignment = CENTER_ALIGN
    cell_b.border = THIN_BORDER

    # Column C: Modal Text
    cell_c = ws.cell(row=row, column=3, value=step.get('modalText', ''))
    cell_c.font = BODY_FONT
    cell_c.alignment = WRAP_ALIGN
    cell_c.border = THIN_BORDER

    # Column D: Voiceover Script
    cell_d = ws.cell(row=row, column=4, value=step.get('voiceover', ''))
    cell_d.font = BODY_FONT
    cell_d.alignment = WRAP_ALIGN
    cell_d.border = THIN_BORDER

    # Column E: Notes
    cell_e = ws.cell(row=row, column=5, value=step.get('notes', ''))
    cell_e.font = BODY_FONT
    cell_e.alignment = WRAP_ALIGN
    cell_e.border = THIN_BORDER

    # Alternating row colors
    if row % 2 == 0:
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = LIGHT_BLUE_FILL


def add_quiz_section(ws, start_row, quiz_data):
    """Add quiz questions with styled A/B/C options."""
    row = start_row

    for q in quiz_data:
        ws.row_dimensions[row].height = 160

        # Column A: Quiz ID
        cell_a = ws.cell(row=row, column=1, value=q.get('id', ''))
        cell_a.font = QUIZ_FONT
        cell_a.alignment = CENTER_ALIGN
        cell_a.border = THIN_BORDER
        cell_a.fill = QUIZ_FILL

        # Column B: Question with options
        options = q.get('options', {})
        correct = q.get('correct', '')
        question_text = f"{q.get('question', '')}\n\n"
        for opt, text in options.items():
            marker = "[CORRECT] " if opt == correct else ""
            question_text += f"{marker}{opt}) {text}\n"
        cell_b = ws.cell(row=row, column=2, value=question_text.strip())
        cell_b.font = BODY_FONT
        cell_b.alignment = WRAP_ALIGN
        cell_b.border = THIN_BORDER
        cell_b.fill = QUIZ_FILL

        # Column C: Correct answer feedback
        feedback = q.get('feedback', {})
        correct_feedback = f"Correct Answer: {correct}\n\n{feedback.get('correct', '')}"
        cell_c = ws.cell(row=row, column=3, value=correct_feedback)
        cell_c.font = BODY_FONT
        cell_c.alignment = WRAP_ALIGN
        cell_c.border = THIN_BORDER
        cell_c.fill = CORRECT_FILL

        # Column D: Incorrect feedback
        incorrect_feedback = "Incorrect Answer Feedback:\n\n"
        for opt in options.keys():
            if opt != correct:
                fb = feedback.get(opt, '')
                incorrect_feedback += f"{opt}: {fb}\n\n"
        cell_d = ws.cell(row=row, column=4, value=incorrect_feedback.strip())
        cell_d.font = BODY_FONT
        cell_d.alignment = WRAP_ALIGN
        cell_d.border = THIN_BORDER
        cell_d.fill = INCORRECT_FILL

        # Column E: Notes
        cell_e = ws.cell(row=row, column=5, value="Multiple choice quiz question")
        cell_e.font = BODY_FONT
        cell_e.alignment = WRAP_ALIGN
        cell_e.border = THIN_BORDER
        cell_e.fill = QUIZ_FILL

        row += 1

    return row


def create_sheet(wb, section_data, is_first=False):
    """Create a complete sheet with tour steps and quiz."""
    title = section_data.get('sheetName', section_data.get('title', 'Sheet'))
    # Excel sheet name max 31 chars
    if len(title) > 31:
        title = title[:31]

    if is_first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    setup_column_widths(ws)
    add_headers(ws)

    steps = section_data.get('steps', [])
    quiz = section_data.get('quiz', [])

    # Add tour steps
    row = 2
    for step in steps:
        add_tour_step(ws, row, step)
        row += 1

    # Add separator row
    ws.row_dimensions[row].height = 20
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col, value="" if col > 1 else "QUIZ")
        cell.fill = NAVY_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    row += 1

    # Add quiz section
    add_quiz_section(ws, row, quiz)

    return len(steps), len(quiz)


def main():
    print("\n" + "=" * 60)
    print("  MPP Portal eLearning - Storyboard Generator")
    print("=" * 60)

    wb = Workbook()
    total_steps = 0
    total_quiz = 0

    for i, section_file in enumerate(SECTION_FILES):
        filepath = os.path.join(DATA_DIR, section_file)
        if not os.path.exists(filepath):
            print(f"  [SKIP] {section_file} not found")
            continue

        with open(filepath, 'r', encoding='utf-8') as f:
            section_data = json.load(f)

        steps, quiz = create_sheet(wb, section_data, is_first=(i == 0))
        total_steps += steps
        total_quiz += quiz
        print(f"  [OK] {section_data.get('sheetName', section_file)}: "
              f"{steps} steps + {quiz} quiz questions")

    wb.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print(f"  Output: {OUTPUT_FILE}")
    print("=" * 60)
    print(f"\n  Total: {total_steps} steps + {total_quiz} quiz questions")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"\n  Storyboard generated successfully!\n")


if __name__ == "__main__":
    main()
